import openpyxl as xl

name_of_file = input("What is the file name?")

output_file = xl.Workbook()

ws2 = output_file.worksheets[0]

if name_of_file == "":
    name_of_file = "pharmExcel.xlsx"


records = xl.load_workbook(name_of_file, data_only=True)
ws  = records.worksheets[0]

max_rows = ws.max_row
max_columns = ws.max_column


while True:
    column_of_drug = int(input("What column number contains the drug names?"))
    example_rows = 10

    if int(max_rows) < 5:
        example_rows = max_rows
    

    for i in ws.values:
        print(i[column_of_drug-1])
            
    yesno = input("Does this look like the correct row?")

    if yesno in ("y","Y"):
        break
    else:
        continue
        



list_of_drugs = ["Stelara", "Humira", "Zeljanz","Fukitol"]



for iIndex,i in enumerate(list_of_drugs):
    list_of_drugs[iIndex] = i.lower()



print(list_of_drugs)

# the row numbers that contain the drugs we want
accepted_line_numbers = []



new_row = 1
new_column = 1


for iIndex, currentReadRow in enumerate(ws.values):
    
    
    
    if currentReadRow[column_of_drug-1] in list_of_drugs:
        
        
        #print("Drug was found")
        #print(i)

        # for every cell within the row that we are using
        for celldata in currentReadRow:
            #print(f"iIndex is {iIndex}")
            c = ws.cell(row = (iIndex+1),column = new_column)
            #print("*"*10)
            #print(f"Row is {iIndex} and column is {new_column}")
            #print(f"c value is {c.value}")
            #print(c)
            #print("*"*10)
            
            #the output excel workbook is started here
            ws2.cell(row = new_row, column = new_column).value = c.value
            new_column +=1
            
            
        new_column = 1
        
        accepted_line_numbers.append(iIndex+1)
        new_row += 1


yes_no = input("Do you wanna see list of found data?")

if yes_no in ("Yes", "Y", "y", "yes"):
    for i in ws2.values:
        print(i)


    print(f"Drug was found in the following rows: {accepted_line_numbers}")

output_file.save("outputFile.xlsx")
